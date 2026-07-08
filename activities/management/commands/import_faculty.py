from openpyxl import load_workbook

from django.core.management.base import BaseCommand

from activities.models import Faculty


class Command(BaseCommand):

    help = "Import Faculty"


    def handle(self, *args, **kwargs):

        wb = load_workbook("IT Faculty Info.xlsx")

        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):

            faculty_id = str(row[0]).strip()

            faculty_name = str(row[1]).strip()

            Faculty.objects.get_or_create(

                faculty_id=faculty_id,

                faculty_name=faculty_name

            )

        self.stdout.write("Faculty Imported Successfully")